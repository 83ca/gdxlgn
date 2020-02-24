import csv
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.styles import PatternFill


def csv_to_list(csv_path):
    with open(csv_path) as input_csv:
        data_list = [row for row in csv.reader(input_csv)]
    return data_list


class OutputDataSelector:
    def __init__(self, csv_path):
        LABEL_ROW = 0
        LABEL_COLUMN = 0
        DATA_POSITION_COLUMN_MIN = 1
        DATA_POSITION_COLUMN_MAX = 2
        DATA_POSITION_ROW = 0
        data_list = csv_to_list(csv_path)
        # values type: str
        self.label = data_list[LABEL_ROW][LABEL_COLUMN]
        self.data_cols_range = (data_list[DATA_POSITION_ROW][DATA_POSITION_COLUMN_MIN:DATA_POSITION_COLUMN_MAX+1])
        self.selectors = []
        for row in data_list[LABEL_ROW+1:]:
            self.selectors.append(tuple([i for i in row if i != '']))


class XlObjCreator:
    def __init__(self, xl_path):
        self.path = xl_path
        self.wb = load_workbook(self.path)
        self.ws0 = self.wb.active
        self.max_rows = len(tuple(self.ws0.rows))
        self.max_cols = len(tuple(self.ws0.columns))


class XlChartGenerator:
    def __init__(self, xlsx_path, label_row, data_sheet_title=''):
        self.xl = XlObjCreator(xlsx_path)
        self.data_sheet = self.set_data_sheet(data_sheet_title)
        self.label_row = label_row

    def set_data_sheet(self, data_sheet_title):
        if not data_sheet_title:
            data_sheet = self.xl.wb.active
        else:
            data_sheet = self.xl.wb[data_sheet_title]
        return data_sheet

    def load_selector(self, selector_csv_path):
        selector = OutputDataSelector(selector_csv_path)
        xl_labels = [i.value for i in self.data_sheet[self.label_row]]
        selector_index = xl_labels.index(selector.label)
        selector_index_letter = get_column_letter(selector_index+1)

        xl_selector_col = self.data_sheet[selector_index_letter]
        xl_selector_col_str = [str(i.value) for i in xl_selector_col]

        row_selected = []
        for select_set in selector.selectors:
            data_set = []
            for i in select_set:
                i_index = xl_selector_col_str.index(i)
                data_set.append(xl_selector_col[i_index].row)
            row_selected.append(data_set)

        data_column_range = [column_index_from_string(i) for i in selector.data_cols_range]

        return [data_column_range, row_selected]

    def create_chart_selected(self, output_sheet_name, selector_csv_path):
        output_ws = self.xl.wb.create_sheet(output_sheet_name)
        col_range, rows_selected = self.load_selector(selector_csv_path)
        min_col, max_col = col_range
        xvalues = Reference(self.data_sheet, min_col=min_col, min_row=self.label_row, max_col=max_col)

        # 4 cells
        CHART_WIDTH_CELL = 4
        CHART_WIDTH = 1.9 * CHART_WIDTH_CELL
        CHART_HEIGHT = 1.9 * CHART_WIDTH_CELL * 3/4
        CHART_HEIGHT_CELL = int(CHART_HEIGHT/0.5)+2
        CHART_POSITION_ROW_INIT = 2
        CHART_POSITION_COL_INIT = 2
        CHART_REPEAT = 8
        CHART_X_AXIS_TITLE = "X"
        CHART_Y_AXIS_TITLE = "Y"

        row = CHART_POSITION_ROW_INIT
        col = CHART_POSITION_COL_INIT
        r = 0
        for rows in rows_selected:
            chart = ScatterChart()
            id_list = []
            for i in rows:
                min_row = i
                max_row = i

                # ID
                ID_COL_RANGE = [1, 4]
                id0_col_letter = get_column_letter(ID_COL_RANGE[0])
                id1_col_letter = get_column_letter(ID_COL_RANGE[1])
                ids = [j.value for j in self.data_sheet[id0_col_letter + str(min_row): id1_col_letter + str(min_row)][0]]
                if len(ids) > 1:
                    id_title = '-'.join([str(j) for j in ids])
                elif len(ids) == 1:
                    id_title = ids[0]
                else:
                    id_title = 'none'
                id_list.append(id_title)

                values = Reference(self.data_sheet, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)
                # 系列名
                series_title = self.data_sheet[id0_col_letter + str(min_row)].value

                series = Series(values, xvalues, title=series_title)
                chart.series.append(series)

            title_cell = get_column_letter(col) + str(row - 1)
            if len(id_list) > 1:
                chart_title = ', '.join([str(i) for i in id_list])
            elif len(id_list) == 1:
                chart_title = id_list[0]
            else:
                chart_title = ''
            output_ws[title_cell] = chart_title

            # chart style setting
            chart.width = CHART_WIDTH
            chart.height = CHART_HEIGHT
            chart.x_axis.title = CHART_X_AXIS_TITLE
            chart.x_axis.scaling.logBase = 10
            chart.y_axis.title = CHART_Y_AXIS_TITLE
            chart.legend.overlay = True
            chart.legend = None

            output_ws.add_chart(chart, get_column_letter(col)+str(row))

            r += 1
            if r % CHART_REPEAT == 0:
                row = CHART_POSITION_ROW_INIT
                col += CHART_WIDTH_CELL
            else:
                row += CHART_HEIGHT_CELL
        return

    def save(self, path):
        self.xl.wb.save(path)


def main():
    selector_path = "sample/sample_selector.csv"
    xl_path = "sample/sample.xlsx"
    XL_LABEL_ROW = 2
    xl = XlChartGenerator(xl_path, XL_LABEL_ROW)
    xl.create_chart_selected("out", selector_path)
    xl.save("./sample/out.xlsx")


if __name__ == '__main__':
    main()