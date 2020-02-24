import csv
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.chart import ScatterChart, Reference, Series, BarChart


def csv_to_list(csv_path):
    with open(csv_path) as input_csv:
        data_list = [row for row in csv.reader(input_csv)]
    return data_list


class OutputDataSelector:
    def __init__(self, csv_path):
        LABEL_ROW = 0
        LABEL_COLUMN = 0
        DATA_POSITION_COLUMN_MIN = 1
        DATA_POSITION_COLUMN_MAX = 2
        DATA_POSITION_ROW = 0
        data_list = csv_to_list(csv_path)
        # values type: str
        self.label = data_list[LABEL_ROW][LABEL_COLUMN]
        self.data_cols_range = (data_list[DATA_POSITION_ROW][DATA_POSITION_COLUMN_MIN:DATA_POSITION_COLUMN_MAX+1])
        self.selectors = []
        for row in data_list[LABEL_ROW+1:]:
            self.selectors.append(tuple([i for i in row if i != '']))


class XlObjCreator:
    def __init__(self, xl_path):
        self.path = xl_path
        self.wb = load_workbook(self.path)
        self.ws0 = self.wb.active
        self.max_rows = len(tuple(self.ws0.rows))
        self.max_cols = len(tuple(self.ws0.columns))

XL_LABEL_ROW = 2

class XlDataController:
    def __init__(self, xlsx_path, data_sheet_title=''):
        self.xl = XlObjCreator(xlsx_path)
        self.data_sheet = self.set_data_sheet(data_sheet_title)

    def set_data_sheet(self, data_sheet_title):
        if not data_sheet_title:
            data_sheet = self.xl.wb.active
        else:
            data_sheet = self.xl.wb[data_sheet_title]
        return data_sheet

    def load_selector(self, selector_csv_path):
        selector = OutputDataSelector(selector_csv_path)
        # excel row(1, 2, 3, ...)
        # list index(0, 1, 2, 3,...) XL_LABEL_ROW+1-1
        XL_DATA_ROW_INDEX = XL_LABEL_ROW

        xl_labels = [i.value for i in self.data_sheet[XL_LABEL_ROW]]
        selector_index = xl_labels.index(selector.label)
        selector_index_letter = get_column_letter(selector_index+1)

        xl_selector_col = self.data_sheet[selector_index_letter]
        xl_selector_col_str = [str(i.value) for i in xl_selector_col]

        row_selected = []
        for select_set in selector.selectors:
            data_set = []
            for i in select_set:
                i_index = xl_selector_col_str.index(i)
                data_set.append(xl_selector_col[i_index].row)
            row_selected.append(data_set)

        data_column_range = [column_index_from_string(i) for i in selector.data_cols_range]

        return [data_column_range, row_selected]

    def create_chart_selected(self, output_sheet_name, selector_csv_path):
        output_ws = self.xl.wb.create_sheet(output_sheet_name)
        col_range, rows_selected = self.load_selector(selector_csv_path)
        min_col, max_col = col_range
        xvalues = Reference(self.data_sheet, min_col=min_col, min_row=XL_LABEL_ROW, max_col=max_col)
        for rows in rows_selected:
            chart = ScatterChart()
            for i in rows:
                min_row = i
                max_row = i
                values = Reference(self.data_sheet, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)
                series = Series(values, xvalues)
                chart.series.append(series)
            output_ws.add_chart(chart, "B2")

    def save(self, path):
        self.xl.wb.save(path)


def main():
    selector_path = "../sample/sample_selector.csv"
    selector = OutputDataSelector(selector_path)
    print(selector.label)
    print(selector.data_cols_range)
    print(selector.selectors)
    xl_path = "../sample/sample2.xlsx"
    xl = XlDataController(xl_path)
    print(xl.load_selector(selector_path))
    xl.create_chart_selected("out", selector_path)
    xl.save("../sample/out.xlsx")


if __name__ == '__main__':
    main()
